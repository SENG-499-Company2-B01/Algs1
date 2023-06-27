from openpyxl import load_workbook
import json, random

# Courses
wb = load_workbook('courses.xlsx')
ws = wb.active

courses = []
for index, row in enumerate(ws.iter_rows(values_only=True)):
    if index > 0:
        try:
            courses.append(
                {
                    "name": row[1],
                    "shorthand": row[0].replace(" ", ""),
                    "terms_offered": row[2].strip('][').split(', '),
                    "prerequisites": [[]],
                    "corequisites": []
                }
            )
        except:
            print(row)
jsonObject = json.dumps(courses, indent=4)
with open("testing_courses.json", "w") as outfile:
    outfile.write(jsonObject)

# Classrooms
wb = load_workbook('classrooms.xlsx')
ws = wb.active

classrooms = []
for index, row in enumerate(ws.iter_rows(values_only=True)):
    if index > 0:
        try:
            classrooms.append(
                {
                    "capacity": int(row[3]),
                    "building": row[0],
                    "room": row[2],
                    "shorthand": row[1],
                }
            )
        except:
            print(row)
jsonObject = json.dumps(classrooms, indent=4)
with open("testing_classrooms.json", "w") as outfile:
    outfile.write(jsonObject)

# Professors
wb = load_workbook('profs.xlsx')
ws = wb.active

professors = []
for index, row in enumerate(ws.iter_rows(values_only=True)):
    if index > 0:
        try:
            professors.append(
                {
                    "name": f'{row[0]} {row[1]}',
                    "username": f'{row[0]}.{row[1]}',
                    "peng": True,
                    "pref_approved": bool(random.getrandbits(1)),
                    "max_courses": 6,
                    "course_pref": row[3].strip('][').split(', ') if row[3] else [],
                    "unavailable": {}
                }
            )
        except:
            print(row)
jsonObject = json.dumps(professors, indent=4)
with open("testing_profs.json", "w") as outfile:
    outfile.write(jsonObject)